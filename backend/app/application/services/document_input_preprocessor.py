from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from docx import Document
from pypdf import PdfReader
from pypdf.errors import PdfReadError


class DocumentInputPreprocessor:
    def preprocess(self, file_bytes: bytes, filename: str, content_type: str | None) -> tuple[bytes, str]:
        file_kind = self._resolve_kind(filename, content_type)
        if file_kind == "image":
            return file_bytes, ""
        if file_kind == "pdf":
            text = self._extract_pdf_text(file_bytes)
            return self._render_text_to_image(text), text
        if file_kind == "excel":
            text = self._extract_excel_text(file_bytes)
            return self._render_text_to_image(text), text
        if file_kind == "word":
            text = self._extract_word_text(file_bytes, filename)
            return self._render_text_to_image(text), text
        raise ValueError("未対応のファイル形式です")

    @staticmethod
    def _resolve_kind(filename: str, content_type: str | None) -> str:
        ext = Path(filename).suffix.lower()
        if content_type and content_type.startswith("image/"):
            return "image"
        if ext in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
            return "image"
        if content_type == "application/pdf" or ext == ".pdf":
            return "pdf"
        if ext in {".xlsx", ".xlsm", ".xls"}:
            return "excel"
        if ext in {".doc", ".docx"}:
            return "word"
        if content_type in {
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }:
            return "word"
        return "unknown"

    @staticmethod
    def _extract_pdf_text(file_bytes: bytes) -> str:
        try:
            reader = PdfReader(BytesIO(file_bytes))
            texts = []
            for page in reader.pages[:5]:
                texts.append(page.extract_text() or "")
            return "\n".join(texts).strip()
        except PdfReadError as exc:
            raise ValueError("PDFを読み取れません。暗号化または破損の可能性があります") from exc
        except Exception as exc:
            raise ValueError("PDFの処理に失敗しました") from exc

    @staticmethod
    def _extract_excel_text(file_bytes: bytes) -> str:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        lines: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=40, max_col=12, values_only=True):
            cells = [str(cell).strip() for cell in row if cell not in (None, "")]
            if cells:
                lines.append(" | ".join(cells))
        return "\n".join(lines).strip()

    @staticmethod
    def _extract_word_text(file_bytes: bytes, filename: str) -> str:
        ext = Path(filename).suffix.lower()
        if ext == ".docx":
            doc = Document(BytesIO(file_bytes))
            lines = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
            return "\n".join(lines).strip()
        # .doc は構造抽出が難しいため、MVPでは文字列化のみ行う
        return file_bytes.decode("utf-8", errors="ignore").strip()

    @staticmethod
    def _render_text_to_image(text: str) -> bytes:
        safe_text = text or "(no extracted text)"
        clipped = safe_text[:6000]
        image = Image.new("RGB", (1600, 2200), color=(255, 255, 255))
        draw = ImageDraw.Draw(image)
        font = ImageFont.load_default()
        draw.multiline_text((20, 20), clipped, fill=(0, 0, 0), font=font, spacing=4)
        buffer = BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()
